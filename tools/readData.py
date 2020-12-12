import openpyxl
import yaml

from setting import DIR_NAME


class ReadData:
    '''
    列表嵌列表
    '''

    def get_excel(self):
        # wb 工作目录
        wb = openpyxl.load_workbook('./data.xlsx')
        # ws sheet页
        ws = wb['测试用例']
        # 构造数据  [[],[],[]]
        all_cases = []

        selected_data_area = ws.iter_rows(min_row= 2, max_row= ws.max_row, min_col= 1, max_col= ws.max_column)
        print(selected_data_area)
        for rows in selected_data_area:
            cell_valus_list = [cell.value for cell in rows]
            print('每一行的cell值的列表是{}'.format(cell_valus_list))
            all_cases.append(cell_valus_list)

        return all_cases

    def get_yaml(self, filename,key):
        '''
        解析yaml。yml,然后得到列表嵌套字典的数据格式
        读取json的命令：
        json.load()读
        json.dump()写
        读取yaml的命令：
        yaml.safe_load()

        :return:
        '''
        with open(DIR_NAME+'/data/%s.yml' % filename, encoding='utf-8') as f:
            yaml_data = yaml.safe_load(f)
            # print(yaml_data)
            # 构造空列表
            # data_list = []
            cases_dict = yaml_data.get(key)
            case_object = cases_dict.values()
            # 可迭代对象，for in 获取，list（）获取，列表.extend()获取
            # data_list.extend(case_object)
            data = list(case_object)
            return data

    def get_yaml1(self, key):
        '''
        解析yaml。yml,然后得到列表嵌套字典的数据格式
        读取json的命令：
        json.load()读
        json.dump()写
        读取yaml的命令：
        yaml.safe_load()
        :return:
        '''
        with open('../data/login_data.yml', encoding='utf-8') as f:
            yaml_data = yaml.safe_load(f)
            print(yaml_data)
            # 构造空列表
            data_list = []
            cases_dict = yaml_data.get(key)
            case_object = cases_dict.values()
            # 可迭代对象，for in 获取，list（）获取，列表.extend()获取
            # data_list.extend(case_object)
            data = list(case_object)
            return data



if __name__ == "__main__":
    # all_cases = ReadData().get_excel()
    # print(all_cases)
    print(ReadData().get_yaml('login_data', 'test_login'))

